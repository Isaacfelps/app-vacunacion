#!/usr/bin/env python3
"""
Consolidate all anti-rabies vaccination Excel data into a single JSON file.
Reads 2025 municipality-level files and 2026 weekly files.
"""
import os, json, glob, re
from datetime import datetime, timedelta
import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
OUTPUT = os.path.join(BASE, "data.json")

# Columns we care about (emoji-stripped matching)
COLUMN_MAP_2024 = {
    "REGISTRO VACUNACIONID": "id",
    "FECHA/HORA": "fecha",
    "FECHA Y HORA": "fecha",
    "ENTIDAD": "entidad",
    "MUNICIPIO": "municipio",
    "BARRIO": "barrio",
    "VEREDA": "vereda",
    "NOMBRE DEL PROPIETARIO": "nombre_propietario",
    "NUMERO DE IDENTIFICACION": "identificacion",
    "TELEFONO": "telefono",
    "NOMBRE DEL ANIMAL": "nombre_animal",
    "ESPECIE": "especie",
    "EDAD": "edad",
    "GENERO": "genero",
    "VACUNADOR": "vacunador"
}

def strip_emoji(s):
    """Remove emoji and special chars, keep alphanumeric and basic punctuation."""
    if not s:
        return ""
    # Remove common emoji prefixes
    import unicodedata
    cleaned = []
    for ch in str(s):
        cat = unicodedata.category(ch)
        if cat.startswith('So') or cat.startswith('Sk') or cat.startswith('Mn'):
            continue
        cleaned.append(ch)
    return ''.join(cleaned).strip()

def normalize_col(col_name):
    """Normalize column name for matching."""
    if not col_name:
        return ""
    s = strip_emoji(str(col_name)).upper().strip()
    # Remove zero-width joiners and other invisible chars
    s = re.sub(r'[\u200d\ufe0f\u200b]', '', s)
    return s

def find_column_mapping(headers):
    """Map actual header positions to our field names."""
    mapping = {}
    for i, header in enumerate(headers):
        norm = normalize_col(header)
        for key, field in COLUMN_MAP_2024.items():
            if key in norm:
                mapping[i] = field
                break
    return mapping

def parse_date(val):
    """Parse date from various formats."""
    if isinstance(val, datetime):
        return val
    if isinstance(val, str):
        for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"]:
            try:
                return datetime.strptime(val.strip(), fmt)
            except:
                pass
    return None

def compute_estado(fecha_vacuna):
    """Compute vaccination status based on date.
    Anti-rabies vaccine validity: 1 year.
    Verde (VIGENTE): vaccinated within last 10 months
    Amarillo (POR VENCER): between 10-12 months
    Rojo (VENCIDA): more than 12 months ago
    """
    if not fecha_vacuna:
        return "VENCIDA"
    now = datetime(2026, 3, 27)  # Current date
    diff = (now - fecha_vacuna).days
    if diff <= 300:  # ~10 months
        return "VIGENTE"
    elif diff <= 365:  # ~12 months
        return "POR VENCER"
    else:
        return "VENCIDA"

def extract_year(fecha):
    """Extract vaccination year from date."""
    if isinstance(fecha, datetime):
        return fecha.year
    return None

def safe_str(val):
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M")
    if isinstance(val, (int, float)):
        return str(int(val)) if val == int(val) else str(val)
    return str(val).strip()

def process_workbook(filepath, default_municipio=None, default_year=None):
    """Process a single Excel workbook and return list of records."""
    records = []
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        print(f"  ERROR reading {filepath}: {e}")
        return records
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue
        
        # Find header row (first row with recognizable columns)
        header_row = None
        for i, row in enumerate(rows[:5]):
            cols = [normalize_col(c) for c in row if c]
            if any("IDENTIFICACION" in c or "MUNICIPIO" in c or "ESPECIE" in c for c in cols):
                header_row = i
                break
        
        if header_row is None:
            continue
        
        col_map = find_column_mapping(rows[header_row])
        if not col_map:
            continue
        
        for row in rows[header_row + 1:]:
            if not row or all(v is None for v in row):
                continue
            
            record = {}
            for col_idx, field_name in col_map.items():
                if col_idx < len(row):
                    record[field_name] = row[col_idx]
            
            # Skip empty records completely (to prevent data loss, keep if any mapped field is present)
            if not any(v for v in record.values() if v is not None and str(v).strip() != ""):
                continue
            
            # Normalize
            fecha = parse_date(record.get("fecha"))
            municipio = safe_str(record.get("municipio", default_municipio or "")).upper().strip()
            
            # Determine year
            año = None
            if fecha:
                año = fecha.year
                # Some 2024 files have dates in early 2025 – use file context
                if default_year and año != default_year:
                    año = default_year if abs(año - default_year) <= 1 else año
            else:
                año = default_year
            
            # Build final record
            final = {
                "id": safe_str(record.get("id", "")),
                "fecha": fecha.strftime("%Y-%m-%d") if fecha else "",
                "entidad": safe_str(record.get("entidad", "")).upper(),
                "municipio": municipio,
                "barrio": safe_str(record.get("barrio", "")),
                "vereda": safe_str(record.get("vereda", "")),
                "nombre_propietario": safe_str(record.get("nombre_propietario", "")).upper(),
                "identificacion": safe_str(record.get("identificacion", "")),
                "telefono": safe_str(record.get("telefono", "")),
                "nombre_animal": safe_str(record.get("nombre_animal", "")),
                "especie": safe_str(record.get("especie", "")).upper(),
                "edad": safe_str(record.get("edad", "")),
                "genero": safe_str(record.get("genero", "")),
                "vacunador": safe_str(record.get("vacunador", "")),
                "año": año if año else (default_year or 2025),
                "estado": compute_estado(fecha),
            }
            
            records.append(final)
    
    wb.close()
    return records

def main():
    all_records = []
    
    # === Process 2025 data ===
    dir_2025 = os.path.join(BASE, "2025")
    municipio_map = {
        "1.MOCOA": "MOCOA",
        "2.VILLAGARZON": "VILLAGARZÓN",
        "3.PUERTO GUZMAN": "PUERTO GUZMÁN",
        "4.PUERTO ASIS": "PUERTO ASÍS",
        "5.PUERTO CAICEDO": "PUERTO CAICEDO",
        "6.PUERTO LEGUIZAMO": "PUERTO LEGUÍZAMO",
        "7.LA HORMIGA": "VALLE DEL GUAMUEZ",
        "8.ORITO": "ORITO",
        "9.SAN MIGUEL": "SAN MIGUEL",
        "10.COLON": "COLÓN",
        "11.SIBUNDOY": "SIBUNDOY",
        "12.SANTIAGO": "SANTIAGO",
        "13.SAN FRANCISCO": "SAN FRANCISCO",
    }
    
    print("Processing 2025 data...")
    for folder_name, muni_name in municipio_map.items():
        folder_path = os.path.join(dir_2025, folder_name)
        if not os.path.isdir(folder_path):
            print(f"  Folder not found: {folder_path}")
            continue
        xlsx_files = glob.glob(os.path.join(folder_path, "*.xlsx"))
        print(f"  {muni_name}: {len(xlsx_files)} files")
        for f in sorted(xlsx_files):
            recs = process_workbook(f, default_municipio=muni_name, default_year=2025)
            all_records.extend(recs)
    
    # === Process 2026 data ===
    dir_2026 = os.path.join(BASE, "2026")
    print("\nProcessing 2026 data...")
    xlsx_files = sorted(glob.glob(os.path.join(dir_2026, "*.xlsx")))
    print(f"  {len(xlsx_files)} files found")
    for f in xlsx_files:
        print(f"  Processing {os.path.basename(f)}...")
        recs = process_workbook(f, default_year=2026)
        all_records.extend(recs)
    
    # Summary
    print(f"\nTotal records: {len(all_records)}")
    especies = {}
    municipios = {}
    años = {}
    for r in all_records:
        esp = r.get("especie", "OTRO")
        especies[esp] = especies.get(esp, 0) + 1
        mun = r.get("municipio", "DESCONOCIDO")
        municipios[mun] = municipios.get(mun, 0) + 1
        a = r.get("año", "?")
        años[a] = años.get(a, 0) + 1
    
    print(f"Especies: {especies}")
    print(f"Municipios: {municipios}")
    print(f"Años: {años}")
    
    # Process ANIMAL DESCONOCIDO
    unknown_animal_data = []
    unknown_dir = os.path.join(BASE, "Animal desconocido 2026")
    unknown_file = os.path.join(unknown_dir, "ANIMAL DESCONOCIDO.xlsx")
    
    # Get all jpg images in unknown_dir
    unknown_images = []
    if os.path.exists(unknown_dir):
        unknown_images = [f for f in os.listdir(unknown_dir) if f.lower().endswith('.jpg')]

    if os.path.exists(unknown_file):
        try:
            wb_unk = openpyxl.load_workbook(unknown_file, read_only=True, data_only=True)
            ws_unk = wb_unk.active
            rows = list(ws_unk.iter_rows(values_only=True))
            if len(rows) > 1:
                headers = [str(h).strip() if h else "" for h in rows[0]]
                for row in rows[1:]:
                    if not any(row): continue
                    record = {}
                    for i, h in enumerate(headers):
                        if i < len(row):
                            val = row[i]
                            if isinstance(val, datetime):
                                val = val.strftime("%Y-%m-%d")
                            record[h] = safe_str(val) if val is not None else ""
                    
                    # Find matching image
                    animal_id = record.get("ANIMAL_CALLE_ID", "")
                    if animal_id:
                        for img in unknown_images:
                            if str(animal_id) in img:
                                record["foto"] = f"Animal desconocido 2026/{img}"
                                break

                    unknown_animal_data.append(record)
            wb_unk.close()
            print(f"\nANIMAL DESCONOCIDO: {len(unknown_animal_data)} records")
        except Exception as e:
            print(f"Error reading {unknown_file}: {e}")
    
    # Write JSON
    with open(OUTPUT, 'w', encoding='utf-8') as f:
        json.dump(all_records, f, ensure_ascii=False, indent=None)
        
    # Write JS for local file:// protocol access without CORS
    js_output = os.path.join(BASE, "data.js")
    with open(js_output, 'w', encoding='utf-8') as f:
        f.write("window.vaccinationData = ")
        json.dump(all_records, f, ensure_ascii=False, indent=None)
        f.write(";\n")
        f.write("window.unknownAnimalData = ")
        json.dump(unknown_animal_data, f, ensure_ascii=False, indent=None)
        f.write(";\n")
    
    print(f"\nData written to {OUTPUT} and {js_output}")

if __name__ == "__main__":
    main()
